#!/bin/python3

# Copyright 2021 National Technology & Engineering Solutions of Sandia, LLC (NTESS). 
# Under the terms of Contract DE-NA0003525 with NTESS, the U.S. Government retains 
# certain rights in this software.
# 
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
# 
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
# 
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

import os
import sys
import time
import curses
import queue
import readline
import paramiko
import threading
from openpyxl import load_workbook
import configparser
import requests
import urllib3
import json
import http.client as http_client
import re
from datetime import datetime
import time
import traceback

class Scenario_Data(object):
	def __init__(self,XLSX_File):
		self.Scenario = self.read(XLSX_File,'scenario')
		self.Effects = self.read(XLSX_File,'effects')
		self.Logs = self.read(XLSX_File, 'logs')
		self.Scenario_valid = self.Scenario_validate(self.Scenario, self.Effects, self.Logs)

	def Scenario_validate(self,Scenario_data, Effects_data, Logs_Data):
		errors = 0

		#verification of types, structure, etc.
		for data in [Scenario_data, Effects_data, Logs_Data]:
			if not type(data) is dict:
				print("Validator was handed a non-dict object!")
				raise ValueError("Validator was handed a non-dict object!")

			#validate files in scenarios
			for i in data.keys():
				for j in data[i].keys():
					if j == "effects":
						for n in data[i][j]:
							if n != "None" and n != None:
								if not (n in Effects_data.keys()):
									print('Effect refference, {}, does not exist!'.format(n))
									errors  += 1
					if j == "logs":
						for n in data[i][j]:
							if n != "None" and n != None:
								if not (n in Logs_Data.keys()):
									print('Log refference, {}, does not exist!'.format(n))
									errors  += 1
					if j == "scene_children":
						for n in data[i][j]:
							if n != "None" and n != None:
								if not (n in data.keys()):
									print('Scene child, {}, does not exist!'.format(n))
									errors  += 1
					if j[-4:] == "file":
						for n in data[i][j]:
							if n != "None" and n != None:
								if not os.path.exists(n):
									print('{} \nFile does not exist!'.format(n))
									errors  += 1
							elif j == "config_file" and (n == "None" or n == None):
									print('Config file, {}, does not exist!'.format(n))
									errors  += 1

		print('There are {} errors in the Scenario File.'.format(errors))
		return errors

	def read(self, XLSX_File, sheet):
		#reads in the XLSX file and converts it to a dictionary
		wb = load_workbook(filename=XLSX_File)
		
		#raise error if sheet does not exist
		if not (sheet in wb.sheetnames):
			print('Table is missing sheet: {}'.format(sheet))
			raise ValueError('Table is missing sheet: {}'.format(sheet))
		else:
			ws = wb[sheet]
			#grab names of columns for dictionary
			tags = [str(ws['1'][i].value).lower() for i in range(0,ws.max_column)]
			#grab all rows after first and convert into a dict of dict
			d = {}
			for row in list(ws.rows)[1:]:
				rowdict = dict()
				for i in range(1,ws.max_column):
					#split lists on ";" if not the discription
					if tags[i] == "description": 
						rowdict[tags[i]] = str(row[i].value)
					elif tags[i][-8:] == "_command":
						rowdict[tags[i]] = str(row[i].value).split("\\n")
					else:
						rowdict[tags[i]] = [x.strip() for x in str(row[i].value).split(";")]
				d[str(row[0].value)] = rowdict
			wb.close()
			return d

class Log_Controller(object):


	# CONSTRUCTOR

	def __init__(self, Scenario, Log_ID, message_queue, error_queue, index_queue):
		self.Message_Queue = message_queue
		self.Error_message_queue = error_queue
		self.Index_queue = index_queue
		self.Event = threading.Event()
		self.thread = None
		self.Scenario = Scenario
		self.Log_ID = Log_ID
		self.conf_file = Scenario.Logs[Log_ID]['config_file'][0]
		self.log_file = Scenario.Logs[Log_ID]['log_file'][0]
		self.setup(self.conf_file)
		if Scenario.Logs[Log_ID]['log_index'][0] != "None" and Scenario.Logs[Log_ID]['log_index'][0] != None:
			self.index = Scenario.Logs[Log_ID]['log_index'][0]
		self.Index_queue.put(self.index)
		if Scenario.Logs[Log_ID]['log_time'][0] != "None" and Scenario.Logs[Log_ID]['log_time'][0] != None:
			self.time_option = Scenario.Logs[Log_ID]['log_time'][0]
		self.notify("Log Controller created with options:\n\tIP: " + self.ip + ":" + str(self.port) + "\n\tSSL: " + str(self.security) + "\n\tTimestamps: " + self.time_option + "\n\tIndex: " + self.index + "\n\tAuthentication: " + self.username + ":" + self.password)

	# SETUP
	# conf_file : string, filename of .conf file; sample format:
	#	[ELK]
	#	ip = 10.1.7.135
	#	port = 9200
	#	time = 2022-12-13T12:00:00.000Z
	#	username = elastic
	#	password = password
	#	index = test
	#	security = False
	#	delay = False
	def setup(self, conf_file):
		if conf_file != "None" and conf_file != None:
			#self.notify('Setting up log controller object with config: ' + conf_file + ' . . .')

			config = configparser.ConfigParser()
			config.read(conf_file)

			if('ELK' not in config):
				self.error("Bad config provided.")
				return -1		

			options=['delay', 'ip', 'port', 'time', 'index', 'username', 'password', 'security']
			for opt in options:
				if opt not in config['ELK']:
					self.error("Bad config provided. " + opt + " not found.")
					return -1

			if(config['ELK']['delay'] == 'True'):
				self.delay = True
			else:
				self.delay = False
			if(config['ELK']['security'] == 'True'):
				self.security = True
			else:
				self.security = False
			self.ip = config['ELK']['ip']
			self.port = config['ELK']['port']
			self.time_option = config['ELK']['time']
			self.index = config['ELK']['index']
			self.username = config['ELK']['username']
			self.password = config['ELK']['password']

			#self.notify("Log Controller created with options:\n\tIP: " + self.ip + ":" + str(self.port) + "\n\tSSL: " + str(self.security) + "\n\tTimestamps: " + self.time_option + "\n\tIndex: " + self.index + "\n\tAuthentication: " + self.username + ":" + self.password)
		else:
			self.Error_message_queue.put("Missing config file for {}. Log is disabled.".format(self.Log_ID))
			self.Event.set()

	# CONSOLE OUTPUT FUNCTIONS

	#print to queue with yellow
	def notify(self, message):
		self.Message_Queue.put("[+] {}".format(message))

	#print to queue with red
	def error(self, message):
		self.Error_message_queue.put("[!] {}".format(message))


	# MAIN PROGRAM FUNCTIONS

	#parse json log file into array of strings for manipulation later
	#this is needed because logstash sometimes (always) does not use newlines between logs in a sensible manner
	# my_file : string, logstash output to file using json plugin
	def parse_logs(self, my_file):
		try:
			decoder = json.JSONDecoder()
			fileobj = open(my_file, 'r')

			all_logs = []
			contents = fileobj.readlines()
			for line in contents:
				if self.Event.is_set():
					break

				line=line.strip()
				pos = 0
				while not pos == len(line):
					curr_log, length = decoder.raw_decode(line[pos:])
					all_logs.append(curr_log)
					pos += length

				#edge case where an empty string ends up in there
				while("" in all_logs):
					all_logs.remove("")

			return all_logs
		except Exception as e:
			self.error( "Error parsing " + my_file + "\n" + str(e) )
			return ("failed.")


	#sample timestamp: 2022-12-09T19:14:25.412Z
	#KEY ASSUMPTION = LOGS ARE INGESTED IN ORDER BY TIMESTAMP
	#format (supported) time fields according to input
	# logs : array of strings, output from parse_logs]
	# time_option : string in {no_update, now, <timestamp>} ; no option provided uses provided .conf
	def update_timestamps(self, logs, time_option="default"):

		datetime_format = "%Y-%m-%dT%H:%M:%S.%fZ"
		earliest_timestamp = ""
		earliest_datetime = None

		if time_option == "default":
			time_option = self.time_option

		if time_option == 'no_update':
			return logs
		elif time_option == 'now':
			origin_datetime = datetime.utcnow()
		elif re.match(r'\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}.\d{3}Z', time_option):
			origin_datetime = datetime.strptime(time_option, datetime_format)
		else:
			self.error("Time option <" + time_option + "> not supported. See help for more details. Exiting . . .")
			return(-1)


		#self.notify("Updating timestamps using option: " + time_option)

		new_logs = []
		
		for i in logs:
			if self.Event.is_set():
				break
			#I vehemently oppose this, but couldn't think of a faster way
			log_string = json.dumps(i)
			#finding strings that match format coming out of logstash and update - this supports winlogbeat format and possibly others
			result = re.finditer(r'\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}.\d{3}Z', log_string)
			for j in result:
				if self.Event.is_set():
					break
				if earliest_timestamp == "":
					earliest_timestamp = j.group(0)
				earliest_datetime = datetime.strptime(earliest_timestamp, datetime_format)
				current_datetime = datetime.strptime(j.group(0), datetime_format)
				time_delta = current_datetime - earliest_datetime
				new_datetime = origin_datetime + time_delta
				new_timestamp = new_datetime.strftime(datetime_format)
				new_timestamp = new_timestamp[:-4] + 'Z' #python is automatically printing microseconds, chop off 3 LSDs + the 'Z' (zulu) and re-add the Z
				log_string = log_string.replace(j.group(0), new_timestamp)
			#add @timestamp option in case of "ts" : "<epoch>" - this supports bro log format
			if "ts" in i:
				if isinstance(i["ts"], float):
					if earliest_datetime == None:
						earliest_datetime = datetime.fromtimestamp(i["ts"])
						earliest_timestamp = earliest_datetime.strftime(datetime_format)
					current_datetime = datetime.fromtimestamp(i["ts"])
					time_delta = current_datetime - earliest_datetime
					new_datetime = origin_datetime + time_delta
					new_timestamp = new_datetime.strftime(datetime_format)
					new_timestamp = new_timestamp[:-4] + 'Z' #python is automatically printing microseconds, chop off 3 LSDs + the 'Z' (zulu) and re-add the Z
					i["@timestamp"] = new_timestamp
					i["ts"] = new_datetime.strftime("%s.%f")
					log_string = json.dumps(i)
			if ( len(list(result)) > 0 ) and ("ts" in i):
				self.error("Provided log file has attributes of Zeek log dump and winlogbeat log dump. Exiting to avoid arbitrary behavior. . .")
				return -1
			new_logs.append(json.loads(log_string))
		return new_logs
	
	#forwards logs to elastic instance according to .conf values (ip, port, authentication options, etc.) in one POST
	#if delay is set, will forward on to trickle_logs instead
	# logs : array of strings, output from parse_logs or update_timestamps
	# index : string, no option provided uses provided .conf
	def send_logs(self, logs, index="default"):
		if(index == "default"):
			index = self.index

		urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
		
		if(self.security):
			prot = "https://"
		else:
			prot = "http://"

		es_url = prot + self.ip + ":" + str(self.port) + "/_bulk/?pretty"

		headers = {
			'Content-Type': 'application/json',
		}
		
		actions = []
		for i in logs:
			actions.append("{\"index\": {\"_index\": \"" + index + "\"}}")
			actions.append(json.dumps(i))
		
		body='\n'.join(actions)

		#The bulk request must be terminated by a newline [\\n]
		body += "\n"

		#print(body)
		#return("stdout")
		try:
			response = requests.post(es_url, headers=headers, data=body, auth=(self.username, self.password), verify=False, timeout=5.0)
			return str(response)
		except Exception as e:
			self.error( str(e) )
			return ("failed.")

	#forwards logs to elastic instance according to .conf values (ip, port, authentication options, etc.) in separate POSTs, waits delta of timestamps between POSTs
	# logs : array of strings, output from parse_logs or update_timestamps
	# index : string, no option provided uses provided .conf
	def trickle_logs(self, logs, index="default"):
		if(index == "default"):
			index = self.index

		#self.notify("Trickling logs one by one according to timestamp.")

		datetime_format = "%Y-%m-%dT%H:%M:%S.%fZ"
		latest_datetime = None

		#get logs sorted by time
		logs.sort(key=lambda x: x["@timestamp"])

		#TODO:: if we want total fidelity we can take timestamps during the loop and then diff real time past with time delta in logs.... I assume this time is negligble for now.
		num = 0
		#print(num)
		for i in logs:
			if self.Event.is_set():
				return("thread killed.")
			num += 1
			#body = "{\"index\": {\"_index\": \"" + index + "\"}}"
			#body  + = "\n" + json.dumps(i) + "\n"
			current_datetime = datetime.strptime(i["@timestamp"], datetime_format)
			if (latest_datetime == None):
				latest_datetime = current_datetime
			delta = current_datetime - latest_datetime
			#print("current: ", current_datetime)
			#print("latest: ", latest_datetime)
			#print("delta: ", delta.total_seconds() )
			if(current_datetime > latest_datetime):
				latest_datetime = current_datetime
			self.Event.wait(timeout=delta.total_seconds())
			logslist = [] #the list will only ever have one, but sendlogs wants a list, so we shall oblige
			logslist.append(i)
			response = self.send_logs(logslist, index)
			#print('\033[F' + str(num) )
			if(response != "<Response [200]>"):
				self.error("Bad response log #: " + str(num))

		return ("done.")
	
	#deletes all contents of provided index
	# index : string, no option provided clears index present in .conf
	def clear_index(self, index="default"):
		if(index == "default"):
			index = self.index

		self.notify("clearing index: " + index)

		if(self.security):
			prot = "https://"
		else:
			prot = "http://"

		es_url = prot + self.ip + ":" + str(self.port) + "/" + index + "?pretty"

		headers = {
			'Content-Type': 'application/json',
		}

		response = requests.delete(es_url, headers=headers, data=None, auth=(self.username, self.password), verify=False, timeout=1.0)

		return str(response)
	
	#creating a single function to perform all tasks to be easily called as a thread
	def parse_update_and_send(self, log_file, time_option="default", index="default"):
		self.notify( "parsing logs from: " + log_file )
		logs = self.parse_logs(log_file)
		if( logs == "failed." or logs == None or logs == "None" ):
			self.error( log_file + ": failed to parse json." )
		self.notify( "updating timestampts from: " + log_file + " (w/ timestamp option " + time_option + ")" )
		logs = self.update_timestamps(logs, time_option)
		if(self.delay == True):
			self.notify( "trickling " + log_file + " into index " + index )
			resp = self.trickle_logs(logs, index)
		else:
			self.notify( "bulk sending " + log_file + " into index " + index )
			resp = self.send_logs(logs, index)
		self.notify(log_file + " : " + resp)
		
		#stop safely
		self.Stop()

		#use to set Kill signal and kill trickling of logs in threads
	def Stop(self):
		#sys.stderr.write("killing thread\n")
		self.Event.set()
		self.thread.join()

	def Run(self):
		thread = threading.Thread(target=self.parse_update_and_send, args=[self.log_file], daemon=True)
		self.thread = thread
		self.thread.start()

	def Clear_Thread(self, index="default"):
		#TODO:: what the hell is happening here with "all"? Does it even parse to the thread?
		if(index == "all"):
			self.notify("Clearing all indexes that have been uploaded during this session . . .")
			seen = []
			while not self.Index_queue.empty():
				current = self.Index_queue.get()
				if(current == "all"):
					self.error("While clearing indexes received keyword \"all\". Skipping to avoid infinite recursion. Do not use index \"all\" in the future.")
					continue
				if not current in seen:
					self.Clear_Thread(current)
					seen.append(current)
		
		thread = threading.Thread(target=self.clear_index, args=[index], daemon=True)
		self.thread = thread
		self.thread.start()




class Effects_Agent(object):
	def __init__(self, Scenario, EFX_ID, q, error):
		self.Scenario = Scenario
		self.EFX_ID = EFX_ID
		self.message_queue = q
		self.Error_message_queue = error
		self.threads = []
		self.Event = threading.Event()
		self.EFX_Commands = self.Scenario.Effects[EFX_ID]['effect_command']
		self.username = Scenario.Effects[EFX_ID]['agent_username']
		self.agent_ip = Scenario.Effects[EFX_ID]['agent_ip']
		self.password = Scenario.Effects[EFX_ID]['agent_password']
		self.scp_files = self.Scenario.Effects[EFX_ID]['effect_file']
		self.scp_file_dest = self.Scenario.Effects[EFX_ID]['effect_file_destination']
		self.file_loc_default = '~/'
		
		#fix agent ip to passwords and usernames
		if len(self.agent_ip) > len(self.username) and (self.username[0] != None or self.username != ''):
			self.username.extend([str(self.username[-1]) for i in range(len(self.agent_ip)-len(self.username))])
		elif self.username == None or self.username == '':
			self.Error_message_queue.put('No username for IP: {}'.format(self.agent_ip[0]))

		if len(self.agent_ip) > len(self.password) and (self.password[0] != None or self.password != ''):
			self.password.extend([str(self.password[-1]) for i in range(len(self.agent_ip)-len(self.password))])
		elif self.password == None or self.password == '':
			self.Error_message_queue.put('No password for IP: {}'.format(self.agent_ip[0]))
		
		#fix mismatches in files and locations
		if len(self.scp_files) > len(self.scp_file_dest) and (self.scp_file_dest[0] != None or self.scp_file_dest != ''):
			self.scp_file_dest.extend([str(self.scp_file_dest[-1]) for i in range(len(self.scp_files)-len(self.scp_file_dest))])
		elif self.scp_file_dest == None or self.scp_file_dest == '':
			self.scp_file_dest = [str(self.file_loc_default) for i in range(len(self.scp_files))]
		

	def Commander(self,ID):
		import select
		import socket
		from scp import SCPClient

		username = self.username[ID]
		agent_ip = self.agent_ip[ID]
		password = self.password[ID]
		#connect to an ssh client and run commands
		ssh = paramiko.SSHClient()
		ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
		
		#setup files for upload
		set_continue = False
			 
		#upload files
		f = 0
		for i in self.scp_files:
			#break if thread is closing
			if self.Event.is_set():
				break
			self.message_queue.put('\nEffect Agent: {}@{} \t File: {} \t Dest: {}\n'.format(username,agent_ip,i,self.scp_file_dest[f]))
			try:
				ssh.connect(agent_ip, username=username, password=password, timeout=2)
				scp = SCPClient(ssh.get_transport())
				set_continue = True
			except:
				self.Error_message_queue.put('Could not connect to SSH on {}@{}'.format(username,agent_ip))
			
			if set_continue:
				scp.put(i, remote_path=self.scp_file_dest[f])
				f += 1
				scp.close()
				ssh.close()

		for i in self.EFX_Commands:
			if self.Event.is_set():
				break

			#set up command to print out the PID of the command (technically the PID of the ssh terminal)
			command = 'echo $$; exec bash -c \'' + i + '\''
			#print out the ip of the agent and what command was run
			
			self.message_queue.put('\nEffect Agent: {}@{} \t Command: {} \n'.format(username,agent_ip,i))
			
			#attempt to connect to the ssh client, start an ssh channel, and continuously output the ssh command output
			try:
				ssh.connect(agent_ip, username=username, password=password, timeout=2)
				transport = ssh.get_transport()
				channel = transport.open_session()
				channel.settimeout(0.0)
				channel.set_combine_stderr(True)
				channel.exec_command(command)
				pid_line = True
				while not self.Event.is_set():
					rl, wl, xl = select.select([channel],[],[])
					if channel in rl:
						try:
							message = channel.recv(1024)
							if len(message) == 0:
								break
							#if this is the first line, its the PID
							if pid_line:
								self.message_queue.put('{}@{} => PID: {}'.format(username,agent_ip,message.decode('ascii').strip("\n")))
								pid_line = False
							else:
								self.message_queue.put('{}@{} => {}'.format(username,agent_ip,message.decode('ascii').strip("\n")))
						except socket.timeout:
							pass
				ssh.close()
			except:
				self.Error_message_queue.put('Could not connect to SSH on {}@{}'.format(username,agent_ip))
		#stop safely
		self.Stop()
		
	def Run(self):
		#run the EFX Commmander for each ip in a different thread.
		for i in range(len(self.agent_ip)):
			thread = threading.Thread(target=self.Commander, args=[i], daemon=True)
			self.threads.append(thread)
			thread.start()

	def Wait(self):
		for i in range(len(self.threads)):
			self.threads[i].join()

	def Stop(self):
		self.Event.set()
		for thread in self.threads:
			thread.join()



class Scenario_Engine(object):
	def __init__(self, Scenario):
		self.Scenario = Scenario
		self.current_scene = '0'
		self.scene_children = self.Scenario.Scenario['0']['scene_children']
		self.Log_Controller_thread = []
		self.EFX_Commander_thread = []
		self.Context_thread = []
		self.cli_columns = os.get_terminal_size().columns
		self.cli_lines = os.get_terminal_size().lines
		self.cli_line_top = int(0.2*self.cli_lines)
		self.cli_line_middle = int(0.6*self.cli_lines)
		self.cli_line_bottom = int(0.2*self.cli_lines)
		self.title = 'Sandia Experiment Control System'
		self.title_center = int(self.cli_columns*0.5 - len(self.title)*0.5)
		self.disc_padding = int(3+len("Description:"))
		self.desc_len = int(self.cli_columns-(2*self.disc_padding))
		self.shutdown = False
		self.Sys_Message = None
		self.EFX_message_queue = queue.Queue()
		self.Log_message_queue = queue.Queue()
		self.Context_message_queue = queue.Queue()
		self.Error_message_queue = queue.Queue()
		self.Index_queue = queue.Queue()
		self.scroll_pos = self.cli_line_bottom
		self.scroll_height = 100
		self.sys_log_name = 'ECS_Log'
		self.sys_log_path = './' + self.sys_log_name + '.txt'
	
	def text_wrangler(self, pad, text, columns, rows, x, y, idx=0):
		
		#quick checker for the text values, no 0's or negatives
		def size_check(n):
			if n < 1:
				n = 1
			return n
		#check values and fix
		columns = size_check(columns)
		rows = size_check(rows)
		x = size_check(x)
		y = size_check(y)

		#split lines, then print everything out within boundaries
		text_split = text.splitlines()
		r = 0
		idx_r = 1
		remainder = 0
		for n in text_split:
			text_rows = int(len(n)//columns + (len(n) % (columns) > 0))
			text_rows = size_check(text_rows)
			for i in range(text_rows):
				if idx_r > idx:	
					pad.addstr(y+r,x,n[i*columns:(i+1)*columns])
					r += 1
					if r > rows:
						remainder = idx+r
						break
				else:
					idx_r += 1
			if r > rows:
					break
		return remainder


	def Top_clr(self, Top_pad):
		Top_pad.clear()
		Top_pad.border(0)
		Top_pad.addstr(1,self.title_center,self.title,curses.A_UNDERLINE)
		Top_pad.addstr(3,2,"Current Scene: {}".format(self.current_scene),curses.A_BOLD)
		Scene_child_text = "Scene Children: {}".format(str(self.scene_children).strip("[]"))
		Top_pad.addstr(3,self.cli_columns-(2+len(Scene_child_text)),Scene_child_text,curses.A_BOLD)
		Top_pad.addstr(5,2,"Description:",curses.A_BOLD)

		#print out the description
		desc_text = self.Scenario.Scenario[self.current_scene]['description']
		self.text_wrangler( Top_pad, desc_text, self.desc_len, self.cli_line_top-5, self.disc_padding, 5)
		Top_pad.refresh(0,0,0,0,self.cli_line_top,self.cli_columns-1)

	def mid_clr(self, Middle_pad):
		Middle_pad.clear()
		Middle_pad.scrollok(1)
		Middle_pad.refresh(0,0,self.cli_line_top+1,0,self.cli_line_middle+self.cli_line_top-1,self.cli_columns-1)
	
	def bot_clr(self, Bottom_pad):
		#define a function for clearing the bottom
		def bot_wipe():
			Bottom_pad.clear()
			Bottom_pad.border(0)
			opt_string = "<i>Input Scene ID <Q>Exit <C>Clear <L>List <E>Stop EFX <S>Stop Logs <X>:Clear Index"
			centering_pad = int(((self.cli_columns - len(opt_string))/2)-1)
			Bottom_pad.addstr(self.cli_line_bottom-2,centering_pad,opt_string,curses.A_STANDOUT)
		
		bot_wipe()

		if self.Sys_Message != None:
			rem = self.text_wrangler( Bottom_pad, self.Sys_Message, self.cli_columns-2, self.cli_line_bottom-3, 1, 1)
			while rem > 0:
				Bottom_pad.refresh(0,0,self.cli_line_middle+self.cli_line_top,0,self.cli_lines-1,self.cli_columns-1)
				curses.noecho()
				get_key = Bottom_pad.getch(1,1)
				bot_wipe()
				rem = self.text_wrangler( Bottom_pad, self.Sys_Message, self.cli_columns-2, self.cli_line_bottom-3, 1, 1, rem)
		Bottom_pad.refresh(0,0,self.cli_line_middle+self.cli_line_top,0,self.cli_lines-1,self.cli_columns-1)

	def mid_update_thread(self, Mid_pad, Top_pad, Bottom_pad):

		curses.init_pair(curses.COLOR_GREEN,curses.COLOR_GREEN,curses.COLOR_BLACK)
		curses.init_pair(curses.COLOR_BLUE,curses.COLOR_BLUE,curses.COLOR_BLACK)
		curses.init_pair(curses.COLOR_YELLOW,curses.COLOR_YELLOW,curses.COLOR_BLACK)
		curses.init_pair(curses.COLOR_RED,curses.COLOR_RED,curses.COLOR_BLACK)
		
		#Check if system log exists and create a new file if it exists already.
		n = 1
		while os.path.exists(self.sys_log_path):
			self.sys_log_path = './' + self.sys_log_name + str(n) + '.txt'
			n += 1
		self.Error_message_queue.put("[!] Starting log file @ " + self.sys_log_path)
		
		system_log = open(self.sys_log_path, 'w')

		#define the method to print things to the screen
		def Message_Printer(Message,color):
			system_log.write(Message)
			system_log.write('\n')
			for i in Message.split('\n'):
				if len(i) > (self.cli_columns - 3):
					for n in range(int(len(i)/(self.cli_columns - 3)) + (len(i) % (self.cli_columns - 3) > 0)):
						m = i[n*(self.cli_columns - 3):(n+1)*(self.cli_columns - 3)]
						Mid_pad.addstr(self.cli_line_middle+self.cli_line_top-1,2,m,curses.color_pair(color))
						Mid_pad.scroll(1)
				else:
					Mid_pad.addstr(self.cli_line_middle+self.cli_line_top-1,2,i,curses.color_pair(color))
					Mid_pad.scroll(1)
			Mid_pad.scroll(1)
			Mid_pad.refresh(self.scroll_pos,0,self.cli_line_top+1,0,self.cli_line_middle+self.cli_line_top-1,self.cli_columns-1)
		
		#start loop to capture messages from the queues and print them
		while self.shutdown == False:
			tag = False
			while not self.EFX_message_queue.empty():
				Message_Printer(self.EFX_message_queue.get(),curses.COLOR_GREEN)
				tag = True

			while not self.Log_message_queue.empty():
				Message_Printer(self.Log_message_queue.get(),curses.COLOR_YELLOW)
				tag = True
			
			while not self.Context_message_queue.empty():
				Message_Printer(self.Context_message_queue.get(),curses.COLOR_BLUE)
				tag = True

			while not self.Error_message_queue.empty():
				Message_Printer(self.Error_message_queue.get(),curses.COLOR_RED)
				tag = True

			if tag:
				tag = False
				self.Top_clr(Top_pad)
				self.bot_clr(Bottom_pad)
		
		system_log.close()

	def trash_man(self):
		#I'm the trashman, I clean up dead threads
		
		while self.shutdown == False:
			
			for i in range(10):
				if self.shutdown == False:
					time.sleep(0.5)
				else:
					break

			#clean up EFX trash
			EFX_trash = []
			for idx, EFX in enumerate(self.EFX_Commander_thread):
				if EFX.Event.is_set(): #find trash
					EFX_trash.append(idx)
			for idx in sorted(EFX_trash, reverse=True):
				del self.EFX_Commander_thread[idx] #throw trash around the ring
			
			#clean up Log trash
			Log_trash = []
			for idx, LOG in enumerate(self.Log_Controller_thread):
				if LOG.Event.is_set():
					Log_trash.append(idx)
			for idx in sorted(Log_trash, reverse=True):
				del self.Log_Controller_thread[idx]


	def CLI(self):
		#user interface system
	
		#redirect stderr to file
		sys.stderr = open('./stderr.log', 'w')
	
		#init screen
		stdscr = curses.initscr()
		stdscr.clear()
		stdscr.refresh()
		curses.start_color()
		stdscr.leaveok(True)
		stdscr.keypad(True)

		#generate pads
		Top_pad = curses.newpad(self.cli_line_top,self.cli_columns)
		Middle_pad = curses.newpad(self.scroll_height,self.cli_columns)
		Bottom_pad = curses.newpad(self.cli_line_bottom,self.cli_columns)
		
		#clear out the screen
		self.Top_clr(Top_pad)
		self.mid_clr(Middle_pad)
		self.bot_clr(Bottom_pad)

		#start thread to update log output
		mid_thread = threading.Thread(target=self.mid_update_thread,args=[Middle_pad, Top_pad, Bottom_pad], daemon=True)
		mid_thread.start()

		#trashman thread start
		trash_man_thread = threading.Thread(target=self.trash_man, daemon=True)
		trash_man_thread.start()

		#Start defining function calls that the user can make
		def Exit():
			self.Sys_Message = "Are you sure you want to EXIT? (y/N)"
			self.bot_clr(Bottom_pad)

			curses.echo()
			selection = stdscr.getstr(self.cli_lines-3,1).decode(encoding="utf-8")

			if selection.lower() == 'y' or selection.lower() == 'yes':
				self.shutdown = True
				for i in self.EFX_Commander_thread:
					i.Stop()
				for i in self.Log_Controller_thread:
					i.Stop()
			else:
				self.Sys_Message = None
				curses.noecho()
				self.bot_clr(Bottom_pad)
		
		def Clear():
			self.Sys_Message = None
			self.Top_clr(Top_pad)
			#self.mid_clr(Middle_pad) #removed because I think its unecessary to clear the middle pad?
			self.bot_clr(Bottom_pad)

		def List():
			self.Sys_Message = 'List of all Scene Options: \n{}\n\nList of all Effects Options: \n{}'.format(str(list(self.Scenario.Scenario.keys())).strip("[]"),str(list(self.Scenario.Effects.keys())).strip("[]"))
			self.bot_clr(Bottom_pad)

		def Kill_EFX():
			Efx_threads = [str(i.EFX_ID) for i in self.EFX_Commander_thread]
			Efx_threads_low = [i.lower() for i in Efx_threads]
			if len(Efx_threads) != 0:
				self.Sys_Message = "Select EFX Threads to kill (or all): \n {}".format(str(Efx_threads).strip("[]"))
				self.bot_clr(Bottom_pad)

				curses.echo()
				selection = stdscr.getstr(self.cli_lines-3,1).decode(encoding="utf-8")
				
				if selection.lower() == 'all':
					self.Sys_Message = "Ending Effects"
					self.bot_clr(Bottom_pad)
					for i in self.EFX_Commander_thread:
						i.Stop()
					self.EFX_Commander_thread = []
					self.bot_clr(Bottom_pad)
				elif selection.lower() in Efx_threads_low:
					death_note = [i for i, e in enumerate(Efx_threads_low) if e == selection.lower() ]
					for i in death_note:
						self.EFX_Commander_thread[i].Stop()
					for idx in sorted(death_note, reverse=True):
						del self.EFX_Commander_thread[idx]
					self.Sys_Message = "Killing EFX: {}".format(selection)	
					self.bot_clr(Bottom_pad)

			else:
				self.Sys_Message = "No EFX Threads to kill"
				self.bot_clr(Bottom_pad)

		def Kill_Log_Controller():
			Log_threads = [str(i.Log_ID) for i in self.Log_Controller_thread]
			Log_threads_low = [i.lower() for i in Log_threads]
			if len(Log_threads) != 0:
				self.Sys_Message = "Select Log Threads to kill (or all): \n {}".format(str(Log_threads).strip("[]"))
				self.bot_clr(Bottom_pad)

				curses.echo()
				selection = stdscr.getstr(self.cli_lines-3,1).decode(encoding="utf-8")
				
				if selection.lower() == 'all':
					self.Sys_Message = "Ending Logs"
					self.bot_clr(Bottom_pad)
					for i in self.Log_Controller_thread:
						i.Stop()
					self.Log_Controller_thread = []
					self.bot_clr(Bottom_pad)
				elif selection.lower() in Log_threads_low:
					death_note = [i for i, e in enumerate(Log_threads_low) if e == selection.lower() ]
					for i in death_note:
						self.Log_Controller_thread[i].Stop()
					for idx in sorted(death_note, reverse=True):
						del self.Log_Controller_thread[idx]
					self.Sys_Message = "Killing Logs: {}".format(selection)	
					self.bot_clr(Bottom_pad)

			else:
				self.Sys_Message = "No Log Threads to kill"
				self.bot_clr(Bottom_pad)

		#we have to convert the names of scenario IDs to all lower case and zip into a dict
		#this is so we dont have case-sensitivity with input IDs cause its super annoying
		Scenario_keys = self.Scenario.Scenario.keys()
		Lower_keys = [x.lower() for x in Scenario_keys]
		Selection_keys = dict(zip(Lower_keys,Scenario_keys))

		def index_select():
			self.Sys_Message = "Input Scene ID"
			self.bot_clr(Bottom_pad)

			curses.echo()
			selection = stdscr.getstr(self.cli_lines-3,1).decode(encoding="utf-8")
			
			if selection.lower() in Selection_keys.keys():
				self.current_scene = Selection_keys[selection.lower()]
				self.scene_children = self.Scenario.Scenario[self.current_scene]['scene_children']
				self.Top_clr(Top_pad)
				self.bot_clr(Bottom_pad)
				#EFX - Parsing for effects related to scene and creating threads based on them
				for i in self.Scenario.Scenario[self.current_scene]['effects']:
					if not (i == None or i == 'None'):
						#run EFX threads
						self.EFX_Commander_thread.append(Effects_Agent(self.Scenario,i,self.EFX_message_queue, self.Error_message_queue))
						self.EFX_Commander_thread[-1].Run()
						#holding space for running Log threads and Context threads
				#LOG - Grab log files and send in separate threads
				for i in self.Scenario.Scenario[self.current_scene]['logs']:
					if not (i == None or i == 'None'):
						#send accoring to Log Controller configuration, can change time_option, index if needed later
						self.Log_Controller_thread.append(Log_Controller(self.Scenario,i,self.Log_message_queue, self.Error_message_queue, self.Index_queue))
						self.Log_Controller_thread[-1].Run()
				#TODO::CONTEXT
			else:
					self.Sys_Message = "Not an option try again."
					self.Top_clr(Top_pad)
					self.bot_clr(Bottom_pad)
			
			self.Sys_Message = None
			self.bot_clr(Bottom_pad)

		def clear_index():
			self.Sys_Message = "Input index to clear"
			self.bot_clr(Bottom_pad)

			curses.echo()
			selection = stdscr.getstr(self.cli_lines-3,1).decode(encoding="utf-8")
			if selection != None and selection != "":
				self.Log_Controller_thread.append(Log_Controller(self.Scenario,list(self.Scenario.Logs.keys())[0],self.Log_message_queue, self.Error_message_queue, self.Index_queue))
				self.Log_Controller_thread[-1].Clear_Thread(selection.split()[0])
			
			self.Sys_Message = None
			self.bot_clr(Bottom_pad)

		#define the options and attach to keys
		Options_keys = { 'i':index_select, 'c':Clear, 'l':List, 'e':Kill_EFX, 's':Kill_Log_Controller, 'x':clear_index, 'q':Exit }


		try:
			while self.shutdown == False:
				
				curses.noecho()
				#get_key = stdscr.getch(self.cli_lines-3,1)
				get_key = Bottom_pad.getch(1,1)
				self.Sys_Message = None
				self.bot_clr(Bottom_pad)

				if get_key == curses.KEY_DOWN and self.scroll_pos < self.scroll_height - self.cli_line_middle:
					self.scroll_pos += 1
					Middle_pad.refresh(self.scroll_pos,0,self.cli_line_top+1,0,self.cli_line_middle+self.cli_line_top-1,self.cli_columns-1)

				if get_key == curses.KEY_UP and self.scroll_pos > 0:
					self.scroll_pos -= 1
					Middle_pad.refresh(self.scroll_pos,0,self.cli_line_top+1,0,self.cli_line_middle+self.cli_line_top-1,self.cli_columns-1)

				if chr(get_key) in Options_keys:
					Options_keys[chr(get_key)]()

							
		except Exception as e:
			sys.stderr.write(str(e))
			sys.stderr.write(traceback.format_exc())
			curses.endwin()
			self.shutdown = True

		mid_thread.join()
		trash_man_thread.join()
		curses.endwin()
	
#handy input prompter with prefilled value
def rlinput(prompt, prefill=''):
	readline.set_startup_hook(lambda: readline.insert_text(prefill))
	try:
		return input(prompt)
	finally:
		readline.set_startup_hook()

if __name__ == "__main__":
	#main program
	#check if on windows, suggest fix
	if os.name == 'nt':
		print("Windows sucks, get a better operating system...")
		sys.exit("Bad OS")
	#check for xlsx files around CWD
	files = [f for f in os.listdir('.') if os.path.isfile(f) and f.endswith('.xlsx')]
	#ask for input on which file to use
	Scenario_file = rlinput("Enter Scenario File: ", files[0])
	
	#make sure the file exists, is readable, and has the right extension
	check = False
	if os.path.exists(Scenario_file) and os.access(Scenario_file, os.R_OK) and Scenario_file.endswith('.xlsx'):
		check = True
		try:
			Scenario = Scenario_Data(Scenario_file)
		except Exception as e:
			print("Scenario Data Error: "+ str(e))
			check = False
		if Scenario.Scenario_valid != 0:
				check = False
	#If the file is bad, keep asking user until its right        
	while check == False:
		Scenario_file = rlinput("Try again: ", files[0])
		if os.path.exists(Scenario_file) and os.access(Scenario_file, os.R_OK) and Scenario_file.endswith('.xlsx'):
			check = True
			try:
				Scenario = Scenario_Data(Scenario_file)
			except Exception as e:
				print("Scenario Data Error: "+ str(e))
				check = False
			if Scenario.Scenario_valid != 0:
				check = False

	Engine = Scenario_Engine(Scenario)
	Engine.CLI()

